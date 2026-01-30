import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import sqlite3
import os
import csv
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt

# Matplotlib dark theme ayarları
plt.style.use('dark_background')


class GrafikContainer:
    """Grafik sayfası - DB kayıtlarını grafik ve tablo olarak gösterir"""
    
    def __init__(self, parent_frame, colors):
        self.parent_frame = parent_frame
        self.colors = colors
        self.db_path = 'dosyalar/database.db'
        
        # UI oluştur
        self.setup_ui()
        self.load_data()
        
    def setup_ui(self):
        """UI bileşenlerini oluştur"""
        # Ana container
        main_container = tk.Frame(self.parent_frame, bg=self.colors['bg_dark'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Başlık
        title_frame = tk.Frame(main_container, bg=self.colors['bg_medium'])
        title_frame.pack(fill=tk.X, pady=(0, 20))
        
        title_label = tk.Label(
            title_frame,
            text="📊 Video Kayıtları ve Geçiş Sayımları",
            font=('Segoe UI', 18, 'bold'),
            bg=self.colors['bg_medium'],
            fg=self.colors['text'],
            pady=15
        )
        title_label.pack()
        
        # Yenile butonu
        refresh_btn = tk.Button(
            title_frame,
            text="🔄 Yenile",
            font=('Segoe UI', 10),
            bg=self.colors['accent'],
            fg='white',
            relief=tk.FLAT,
            padx=15,
            pady=8,
            cursor='hand2',
            command=self.load_data
        )
        refresh_btn.pack(side=tk.RIGHT, padx=20)
        self.add_hover_effect(refresh_btn, self.colors['accent'], self.colors['accent_hover'])

        export_btn = tk.Button(
            title_frame,
            text="📤 Excel Dışarıya Aktar",
            font=('Segoe UI', 10),
            bg=self.colors['accent'],
            fg='white',
            relief=tk.FLAT,
            padx=15,
            pady=8,
            cursor='hand2',
            command=self.export_table
        )
        export_btn.pack(side=tk.RIGHT, padx=(0, 10))
        self.add_hover_effect(export_btn, self.colors['accent'], self.colors['accent_hover'])
        
        # İçerik container (yatay: sol %40 tablo, sağ %60 grafik)
        content_frame = tk.Frame(main_container, bg=self.colors['bg_dark'])
        content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Sol panel - Tablo
        left_panel = tk.Frame(content_frame, bg=self.colors['bg_medium'])
        left_panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 10))
        
        table_label = tk.Label(
            left_panel,
            text="📋 Video Kayıtları",
            font=('Segoe UI', 14, 'bold'),
            bg=self.colors['bg_medium'],
            fg=self.colors['text'],
            pady=10
        )
        table_label.pack()
        
        # Tablo için frame
        table_frame = tk.Frame(left_panel, bg=self.colors['bg_medium'])
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        # Treeview (tablo)
        columns = ('ID', 'İsim', 'Tarih', 'Frame Sayısı')
        self.tree = ttk.Treeview(table_frame, columns=columns, show='headings', height=15)
        
        # Sütun başlıkları
        self.tree.heading('ID', text='ID')
        self.tree.heading('İsim', text='Video İsmi')
        self.tree.heading('Tarih', text='Kayıt Tarihi')
        self.tree.heading('Frame Sayısı', text='Frame Sayısı')
        
        # Sütun genişlikleri
        self.tree.column('ID', width=50, anchor=tk.CENTER)
        self.tree.column('İsim', width=200, anchor=tk.W)
        self.tree.column('Tarih', width=150, anchor=tk.CENTER)
        self.tree.column('Frame Sayısı', width=120, anchor=tk.CENTER)
        
        # Scrollbar
        scrollbar_table = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar_table.set)
        
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_table.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Seçim event'i
        self.tree.bind('<<TreeviewSelect>>', self.on_record_select)
        
        # Sağ panel - Grafik
        right_panel = tk.Frame(content_frame, bg=self.colors['bg_medium'])
        right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(10, 0))
        
        graph_label = tk.Label(
            right_panel,
            text="📈 Geçiş Sayımları Grafiği",
            font=('Segoe UI', 14, 'bold'),
            bg=self.colors['bg_medium'],
            fg=self.colors['text'],
            pady=10
        )
        graph_label.pack()
        
        # Grafik için frame
        graph_frame = tk.Frame(right_panel, bg=self.colors['bg_medium'])
        graph_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        
        # Matplotlib figure
        self.fig = Figure(figsize=(8, 6), facecolor=self.colors['bg_medium'])
        self.ax = self.fig.add_subplot(111, facecolor=self.colors['bg_medium'])
        self.ax.tick_params(colors=self.colors['text'])
        self.ax.xaxis.label.set_color(self.colors['text'])
        self.ax.yaxis.label.set_color(self.colors['text'])
        self.ax.title.set_color(self.colors['text'])
        
        # Canvas
        self.canvas = FigureCanvasTkAgg(self.fig, graph_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        
        # Not: "Seçili Kayıt Detayları" alanı kaldırıldı. Seçim sadece grafiği günceller.
        
    def add_hover_effect(self, widget, normal_color, hover_color):
        """Butonlara hover efekti ekle"""
        widget.bind('<Enter>', lambda e: widget.configure(bg=hover_color))
        widget.bind('<Leave>', lambda e: widget.configure(bg=normal_color))
    
    def load_data(self):
        """Veritabanından verileri yükle"""
        if not os.path.exists(self.db_path):
            self.show_empty_state()
            return
        
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Tüm kayıtları getir
            cursor.execute('''
                SELECT id, name, created_at, frame_count
                FROM video_records
                ORDER BY created_at DESC
            ''')
            
            records = cursor.fetchall()
            conn.close()
            
            # Tabloyu temizle ve doldur
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            if not records:
                self.show_empty_state()
                return
            
            for record in records:
                record_id, name, created_at, frame_count = record
                # Tarih formatını düzenle
                date_str = created_at[:19] if created_at else "Bilinmiyor"
                self.tree.insert('', tk.END, values=(record_id, name, date_str, frame_count or 0), tags=(str(record_id),))
            
            # Genel grafik göster
            self.show_overall_graph(records)
            
        except Exception as e:
            print(f"Veri yükleme hatası: {e}")
            self.show_empty_state()
    
    def show_empty_state(self):
        """Boş durum göster"""
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        self.ax.clear()
        self.ax.text(0.5, 0.5, 'Henüz kayıt yok', 
                    ha='center', va='center', 
                    fontsize=14, color=self.colors['text'],
                    transform=self.ax.transAxes)
        self.canvas.draw()
        
        # Detay paneli kaldırıldı
    
    def on_record_select(self, event):
        """Kayıt seçildiğinde detayları göster"""
        selection = self.tree.selection()
        if not selection:
            return
        
        item = self.tree.item(selection[0])
        record_id = item['values'][0]
        
        # Detayları yükle
        self.load_record_details(record_id)
    
    def load_record_details(self, record_id):
        """Seçili kaydın detaylarını yükle"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Geçiş sayımlarını getir
            cursor.execute('''
                SELECT from_area, to_area, count
                FROM transition_counts
                WHERE video_record_id = ?
                ORDER BY count DESC
            ''', (record_id,))
            
            counts = cursor.fetchall()
            
            # Video kaydı bilgisini getir
            cursor.execute('''
                SELECT name FROM video_records WHERE id = ?
            ''', (record_id,))
            record_name = cursor.fetchone()
            record_name = record_name[0] if record_name else "Bilinmeyen"
            
            conn.close()

            if counts:
                # Grafik göster
                self.show_record_graph(counts, record_name)
            else:
                self.ax.clear()
                self.ax.text(0.5, 0.5, f'"{record_name}" için geçiş verisi yok', 
                            ha='center', va='center', 
                            fontsize=12, color=self.colors['text'],
                            transform=self.ax.transAxes)
                self.canvas.draw()
                
        except Exception as e:
            print(f"Detay yükleme hatası: {e}")

    def export_table(self):
        """Seçili kaydı Excel'e (.xlsx) dışarı aktar; geçişleri (kol kol) ayrı sütun yapar.

        Not: openpyxl yoksa CSV olarak kaydeder.
        """
        selection = self.tree.selection()
        if not selection:
            messagebox.showinfo("Bilgi", "Lütfen tablodan bir kayıt seçin, sonra dışarı aktarın.")
            return

        item = self.tree.item(selection[0])
        values = item.get("values", [])
        if not values:
            messagebox.showinfo("Bilgi", "Seçili kayıtta veri bulunamadı.")
            return

        record_id, record_name, record_date, frame_count = values[0], values[1], values[2], values[3]

        # DB'den seçili kaydın geçiş sayımlarını çek
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            cursor.execute('''
                SELECT from_area, to_area, count
                FROM transition_counts
                WHERE video_record_id = ?
                ORDER BY from_area, to_area
            ''', (record_id,))
            counts = cursor.fetchall()
            conn.close()
        except Exception as e:
            messagebox.showerror("Hata", f"Veritabanından geçiş verisi alınamadı:\n{e}")
            return

        # Pivot: her rota (from→to) ayrı sütun
        route_to_count = {f"{from_area} → {to_area}": int(count) for from_area, to_area, count in counts}
        route_columns = sorted(route_to_count.keys())

        base_columns = ["ID", "İsim", "Tarih", "Frame Sayısı"]
        columns = base_columns + (route_columns if route_columns else ["Geçiş Yok"])

        row = [record_id, record_name, record_date, frame_count]
        if route_columns:
            row.extend([route_to_count.get(rc, 0) for rc in route_columns])
        else:
            row.append(0)

        rows = [row]

        # Kullanıcıdan dosya yolu al
        safe_name = "".join(c for c in str(record_name) if c.isalnum() or c in (" ", "-", "_")).strip() or "rapor"
        file_path = filedialog.asksaveasfilename(
            title="Dışarıya Aktar",
            defaultextension=".xlsx",
            filetypes=[("Excel Dosyası", "*.xlsx"), ("CSV Dosyası", "*.csv")],
            initialfile=f"{safe_name}_{record_id}.xlsx"
        )
        if not file_path:
            return

        ext = os.path.splitext(file_path)[1].lower()

        # Excel export (openpyxl)
        if ext == ".xlsx":
            try:
                from openpyxl import Workbook
                from openpyxl.utils import get_column_letter
            except Exception:
                messagebox.showwarning(
                    "Uyarı",
                    "Excel (.xlsx) için 'openpyxl' paketi gerekli. CSV olarak kaydediyorum.\n\n"
                    "Kurulum: pip install openpyxl"
                )
                file_path = os.path.splitext(file_path)[0] + ".csv"
                ext = ".csv"

        try:
            if ext == ".csv":
                with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(columns)
                    writer.writerows(rows)
                messagebox.showinfo("Başarılı", f"CSV kaydedildi:\n{file_path}")
                return

            # .xlsx
            wb = Workbook()
            ws = wb.active
            ws.title = "Seçili Rapor"
            ws.append(columns)
            for r in rows:
                ws.append(list(r))

            # Basit sütun genişliği ayarı
            for col_idx, col_name in enumerate(columns, start=1):
                max_len = max(len(str(col_name)), *(len(str(r[col_idx - 1])) for r in rows if len(r) >= col_idx))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 50)

            wb.save(file_path)
            messagebox.showinfo("Başarılı", f"Excel kaydedildi:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Hata", f"Dışarı aktarma sırasında hata oluştu:\n{e}")
    
    def show_overall_graph(self, records):
        """Genel grafik göster (tüm kayıtlar)"""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # Tüm geçiş sayımlarını topla
            cursor.execute('''
                SELECT from_area, to_area, SUM(count) as total_count
                FROM transition_counts
                GROUP BY from_area, to_area
                ORDER BY total_count DESC
            ''')
            
            all_counts = cursor.fetchall()
            conn.close()
            
            if not all_counts:
                self.ax.clear()
                self.ax.text(0.5, 0.5, 'Genel geçiş verisi yok', 
                            ha='center', va='center', 
                            fontsize=14, color=self.colors['text'],
                            transform=self.ax.transAxes)
                self.canvas.draw()
                return
            
            # Grafik oluştur
            self.ax.clear()
            
            labels = [f"{from_area} → {to_area}" for from_area, to_area, _ in all_counts]
            values = [count for _, _, count in all_counts]
            
            # Bar grafik
            bars = self.ax.bar(range(len(labels)), values, color=self.colors['accent'], alpha=0.7)
            
            # Değerleri üstte göster
            for i, (bar, value) in enumerate(zip(bars, values)):
                height = bar.get_height()
                self.ax.text(bar.get_x() + bar.get_width()/2., height,
                            f'{int(value)}',
                            ha='center', va='bottom', color=self.colors['text'], fontsize=9)
            
            self.ax.set_xlabel('Geçiş Yolları', color=self.colors['text'], fontsize=11)
            self.ax.set_ylabel('Toplam Geçiş Sayısı', color=self.colors['text'], fontsize=11)
            self.ax.set_title('Tüm Kayıtlar - Toplam Geçiş Sayımları', 
                            color=self.colors['text'], fontsize=12, fontweight='bold')
            self.ax.set_xticks(range(len(labels)))
            self.ax.set_xticklabels(labels, rotation=45, ha='right', color=self.colors['text'], fontsize=9)
            self.ax.grid(True, alpha=0.3, color=self.colors['text'])
            
            self.fig.tight_layout()
            self.canvas.draw()
            
        except Exception as e:
            print(f"Grafik oluşturma hatası: {e}")
    
    def show_record_graph(self, counts, record_name):
        """Seçili kayıt için grafik göster"""
        self.ax.clear()
        
        labels = [f"{from_area} → {to_area}" for from_area, to_area, _ in counts]
        values = [count for _, _, count in counts]
        
        # Bar grafik
        bars = self.ax.bar(range(len(labels)), values, color=self.colors['accent'], alpha=0.7)
        
        # Değerleri üstte göster
        for i, (bar, value) in enumerate(zip(bars, values)):
            height = bar.get_height()
            self.ax.text(bar.get_x() + bar.get_width()/2., height,
                        f'{int(value)}',
                        ha='center', va='bottom', color=self.colors['text'], fontsize=9)
        
        self.ax.set_xlabel('Geçiş Yolları', color=self.colors['text'], fontsize=11)
        self.ax.set_ylabel('Geçiş Sayısı', color=self.colors['text'], fontsize=11)
        self.ax.set_title(f'"{record_name}" - Geçiş Sayımları', 
                        color=self.colors['text'], fontsize=12, fontweight='bold')
        self.ax.set_xticks(range(len(labels)))
        self.ax.set_xticklabels(labels, rotation=45, ha='right', color=self.colors['text'], fontsize=9)
        self.ax.grid(True, alpha=0.3, color=self.colors['text'])
        
        self.fig.tight_layout()
        self.canvas.draw()

